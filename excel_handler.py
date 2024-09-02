from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from datetime import datetime
from utils import check_file_exists
from gui import GUI
from constants import OUTPUT_FILE_TYPE, HEADER_NAME_CLOUMNS_ALL, HEADERS_NAME_COLUMNS_TOOLS_LOG
from logger import Logger

class ExcelHandler:
    # Инициализация, объявление всех главных путей до файлов
    def init_values(self, tools_log_path, svod_excel_path, output_path, file_name):
        self.tools_log_path = tools_log_path
        self.svod_excel_path = svod_excel_path
        self.output_path = output_path
        self.file_name = file_name

        self.output_full_path = self.output_path + self.file_name + OUTPUT_FILE_TYPE
        self.logger = Logger()

    # Создание нового и конечного Excel файла обработки
    def define_new_excel_file(self):
        self.wb_new = Workbook()
        self.ws_new = self.wb_new.active
        self.ws_new.title = 'Данные обработки'
        self.ws_new.append(HEADER_NAME_CLOUMNS_ALL)

        self.apply_header_styles()

    # Применение стилей к верхушке заголовков
    def apply_header_styles(self):
        style_fill_head_svod_color = PatternFill(start_color='69fff5', end_color='69fff5', fill_type='solid')
        style_fill_head_tools_log_color = PatternFill(start_color='a8cae4', end_color='a8cae4', fill_type='solid')
        style_align_head = Alignment(vertical='center')
        style_font_head = Font(bold=True)
        style_row_head_height = 45
        for cell in self.ws_new['1:1'[0]]:
            if (cell.value in HEADERS_NAME_COLUMNS_TOOLS_LOG):
                cell.fill = style_fill_head_tools_log_color
            else:
                cell.fill = style_fill_head_svod_color
            cell.alignment = style_align_head
            cell.font = style_font_head
        self.ws_new.row_dimensions[1].height = style_row_head_height

    # Чтение Svod Excel файла
    def define_svod_excel_file(self):
        if check_file_exists(self.svod_excel_path):
            wb_svod = load_workbook(self.svod_excel_path)
            ws_svod = wb_svod['Лист1']
            ws_svod_tuple = tuple(ws_svod.iter_rows(values_only=True))
            self.ws_svod_dict_uid = {}
            for tup in ws_svod_tuple:
                uid_dict = tup[0]
                self.ws_svod_dict_uid[uid_dict] = tup
        else:
            GUI.show_error_messagebox('Ошибка пути', 'Файл Svod Excel не существует по указанному пути')

    # Чтение Tools Log файла
    def define_tools_log_excel_file(self):
        if check_file_exists(self.tools_log_path):
            wb_tools_log = load_workbook(self.tools_log_path)
            self.ws_tools_log = wb_tools_log['Лист1']
        else:
            GUI.show_error_messagebox('Ошибка пути', 'Файл Tools Log не существует по указанному пути')

    def check_if_output_file_exist(self):
        output_path_full = self.output_path + self.file_name + OUTPUT_FILE_TYPE
        if check_file_exists(output_path_full):
            GUI.show_warning_messagebox('Внимание', 'Файл уже существует в папке назначения и будет перезаписан.')

    # Главный обработчик
    def handler(self):
        try:
            start_full_handle = datetime.now()

            self.check_if_output_file_exist()

            st1 = datetime.now()
            st01 = datetime.now()
            self.define_new_excel_file()
            fn01 = datetime.now()
            st02 = datetime.now()
            self.define_svod_excel_file()
            fn02 = datetime.now()
            st03 = datetime.now()
            self.define_tools_log_excel_file()
            fn03 = datetime.now()
            print(str(fn01-st01) + ' - время define_new_excel_file()')
            print(str(fn02-st02) + ' - время define_svod_excel_file()')
            print(str(fn03-st03) + ' - время define_tools_log_excel_file()')
            fn1 = datetime.now()
            print(str(fn1-st1) + ' - время обработки и записи всех файлов в переменные')

            iter: int = 0

            lst_uid = -1
            biggest_date_vidacha = '01.01.1900'
            biggest_date_prihod = '01.01.1900'
            sum_vidano = 0
            count_vidano = 0
            sum_prihod = 0
            count_prihod = 0
            razdel = ''
            type_instrument = ''

            st2 = datetime.now()
            for row in self.ws_tools_log.rows:
                iter += 1

                current_uid = row[0].value
                current_date = row[1].value
                current_vidano = row[3].value
                current_prihod = row[6].value
                current_razdel = row[9].value
                current_type_instrument = row[10].value

                if row[0].value is not None and type(row[0].value) is int:
                    if current_uid != lst_uid:
                        if lst_uid != -1:
                            if biggest_date_vidacha == '01.01.1900':
                                biggest_date_vidacha = ''
                            if biggest_date_prihod == '01.01.1900':
                                biggest_date_prihod = ''

                            obozn = ''
                            naim = ''
                            proizv = ''
                            dop_param = ''
                            ost_irk_20_1 = ''
                            ost_irk_21_1 = ''
                            ost_irk_22_1 = ''
                            ost_irk_23_1 = ''
                            ost_irk_20_2 = ''
                            ost_irk_23_2 = ''
                            stelazh = ''
                            yacheika = ''

                            row_svod_by_uid = self.ws_svod_dict_uid.get(lst_uid)
                            if row_svod_by_uid:
                                obozn = row_svod_by_uid[1]
                                naim = row_svod_by_uid[2]
                                proizv = row_svod_by_uid[3]
                                dop_param = row_svod_by_uid[4]
                                ost_irk_20_1 = row_svod_by_uid[6]
                                ost_irk_21_1 = row_svod_by_uid[8]
                                ost_irk_22_1 = row_svod_by_uid[10]
                                ost_irk_23_1 = row_svod_by_uid[12]
                                ost_irk_20_2 = row_svod_by_uid[14]
                                ost_irk_23_2 = row_svod_by_uid[16]
                                stelazh = row_svod_by_uid[18]
                                yacheika = row_svod_by_uid[19]

                            push_row = [lst_uid, razdel, type_instrument, obozn, naim, proizv, dop_param,
                                        ost_irk_20_1, ost_irk_21_1, ost_irk_22_1, ost_irk_23_1, ost_irk_20_2,
                                        ost_irk_23_2, count_vidano, sum_vidano, biggest_date_vidacha, count_prihod,
                                        sum_prihod, biggest_date_prihod, stelazh, yacheika]
                            self.ws_new.append(push_row)

                        lst_uid = current_uid
                        biggest_date_vidacha = '01.01.1900'
                        biggest_date_prihod = '01.01.1900'
                        sum_vidano = 0
                        count_vidano = 0
                        sum_prihod = 0
                        count_prihod = 0
                        razdel = ''
                        type_instrument = ''

                    razdel = current_razdel
                    type_instrument = current_type_instrument

                    if current_vidano > 0:
                        sum_vidano += current_vidano
                        count_vidano += 1
                        if datetime.strptime(current_date, '%d.%m.%Y') > datetime.strptime(biggest_date_vidacha,
                                                                                           '%d.%m.%Y'):
                            biggest_date_vidacha = current_date

                    if current_prihod > 0:
                        sum_prihod += current_prihod
                        count_prihod += 1
                        if datetime.strptime(current_date, '%d.%m.%Y') > datetime.strptime(biggest_date_prihod,
                                                                                           '%d.%m.%Y'):
                            biggest_date_prihod = current_date
                else:
                    continue
            fn2 = datetime.now()
            print(str(fn2-st2) + ' -> время обработки tools log и записи в новый файл')

            st3 = datetime.now()
            self.stylize_all_handled_cells()
            fn3 = datetime.now()
            print(str(fn3 - st3) + ' -> стилизация каждой ячейки')

            finish_full_handle = datetime.now()
            self.code_work_time = (finish_full_handle - start_full_handle)

            st4 = datetime.now()
            self.create_output_file_excel()
            fn4 = datetime.now()
            print(str(fn4-st4) + ' -> время создания нового файла и конец.')

        except Exception as e:
            GUI.show_error_messagebox('Ошибка!', 'Во время загрузки и обработки файлов произошла ошибка. ' + str(e))
            self.logger.log_error(f'Во время загрузки и обработки файлов произошла ошибка. Путь: {self.output_full_path}.')

    # Стилизация всех ячеек для границ
    def stylize_all_handled_cells(self):
        style_border_header = Border(left=Side(style='medium'),
                                         right=Side(style='medium'),
                                         top=Side(style='medium'),
                                         bottom=Side(style='medium'))
        style_border_all = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
        for row in self.ws_new.rows:
            for cell in row:
                is_auto_size_added = False
                if cell.row != 1:
                    if not is_auto_size_added:
                        self.ws_new.column_dimensions[cell.column_letter].auto_size = True
                        is_auto_size_added = True
                    cell.border = style_border_all
                else:
                    cell.border = style_border_header

        self.ws_new.column_dimensions['B'].width = 30
        self.ws_new.column_dimensions['E'].width = 50

    def create_output_file_excel(self):
        self.wb_new.save(self.output_full_path)

        if check_file_exists(self.output_full_path):
            code_work_time_parse = str(self.code_work_time)[:-3]
            self.logger.log_info(f'Успешно создан файл по пути: {self.output_full_path}. Время выполнения: {code_work_time_parse}')
            GUI.show_info_messagebox('Успех', f'Файл создан по пути: {self.output_full_path}. ' +
                                     f'\nВремя выполнения: {code_work_time_parse}')
        else:
            self.logger.log_error(f'Не удалось создать файл по пути: {self.output_full_path}.')
            GUI.show_error_messagebox('Ошибка', 'Файл не создался по выбранному пути')





# 163871 uid который повторяется по датам 2 раза, но с разным типом инструмента
# нужно брать тот который в svode

# Время загрузки wb_svod =  0:00:17.699499
# Время итераци ws_svod 0:00:01.460051
# Время загрузки wb_tools_log = 0:00:24.833626
# Время полной обработки и создания нового файла = 0:02:55 примерно

"""
    Время отработки приложения на 1000 строк в tools log
    0:00:47.575578 - время обработки и записи всех файлов в переменные
    0:00:02.249947 -> время обработки tools log и записи в новый файл
    0:00:20.009194 -> стилизация каждой ячейки
    0:00:15.283521 -> время создания нового файла и конец.
    
    Время отработки приложения на все 300.000+ строк
    0:00:47.547743 - время обработки и записи всех файлов в переменные
    0:00:07.880961 -> время обработки tools log и записи в новый файл
    0:01:33.022339 -> стилизация каждой ячейки
    0:00:35.260219 -> время создания нового файла и конец.
    
    Оптимизация силей для всех границ время суммарной работы: ~1м. 22с.
    0:00:00.000998 - время define_new_excel_file()
    0:00:20.038660 - время define_svod_excel_file()
    0:00:27.522075 - время define_tools_log_excel_file()
    0:00:47.561733 - время обработки и записи всех файлов в переменные
    0:00:07.921393 -> время обработки tools log и записи в новый файл
    0:00:24.401512 -> стилизация каждой ячейки
    0:00:28.146728 -> время создания нового файла и конец.
"""