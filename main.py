from openpyxl import Workbook, load_workbook
import datetime
import tkinter as tk


# def main():
#     start = datetime.datetime.now()
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Первая листок"
#     ws2 = wb.create_sheet('New Test Лист')
#     ws['A1'] = 5
#     ws['C4'] = 'Записал тут что то'
#     ws.cell(4, 1, 'ws.cell')
#
#     test_data_for_ws2 = [[row*col for col in range(1, 11)] for row in range(1,31)]
#     for row in test_data_for_ws2:
#         ws2.append(row)
#
#     # wb.save('test.xlsx')
#
#
#
#
#     wb2 = load_workbook(r'C:\Users\ilin-mi\Desktop\Копия ToolsLog.xlsx')
#     print(len(list(wb2['Лист1'].values)))
#     finish = datetime.datetime.now()
#     print(finish - start)
#
#
# if __name__ == '__main__':
#     mafr

# from tkinter import *
#
#
# def on_click_btn():
#     txtt = f"Ты нажал..{txt.get('1.0', 'end')}"
#     lbl.configure(text=txtt)
#
# window = Tk()
# window.geometry('400x250')
# window.title("App")
# lbl = Label(window, text="App Lbl", font=("Arial Bold", 20))
# lbl.grid(column=0, row=0)
#
# btn = Button(window, text="Btn",  font=("Arial Bold", 20), bg="green", fg="blue", command=on_click_btn)
# btn.grid(column=4, row=3)
#
# txt = Text(window, width=10, height=1)
# txt.grid(column=1, row=0)
#
#
#
#
# window.mainloop()

import tkinter as tk
from gui import GUI
from excel_handler import ExcelHandler

def main():
    root_window = tk.Tk()
    excel_handler = ExcelHandler()
    gui = GUI(root_window, excel_handler)




    # gui.define_command_handler_files_btn(on_click_run_excel_handler)
    root_window.mainloop()


if __name__ == '__main__':
    main()